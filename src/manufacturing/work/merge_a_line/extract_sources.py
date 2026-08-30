import datetime as dt
import json
from pathlib import Path

from openpyxl import load_workbook

base = Path(r"C:\DEV\manufacturing")
work = base / "work" / "merge_a_line"
files = [
  "2020-01-21_A_Line.xlsx",
  "2020-01-22_A_Line.xlsx",
  "2020-01-23_A_Line.xlsx",
  "2020-01-24_A_Line.xlsx",
  "2020-01-25_A_Line.xlsx",
]
unified_headers = [
  "Date", "Time", "Result", "Periods", "WRITING", "BLE DEVICENAME",
  "BLE MAC ADDRESS", "FCTVER", "MLBSERIAL", "FATPSERIAL", "DSNSERIAL",
  "ETC", "BLE RSSI", "ATIVECURR", "STANBYCURR", "IR/Current",
  "IR LED", "ACC_X", "ACC_Y", "ACC_Z", "원본 파일",
]


def normalize(value, header):
  if value is None:
    return None
  if header == "Date":
    if isinstance(value, (dt.datetime, dt.date)):
      return {"type": "date", "value": value.strftime("%Y-%m-%d")}
    return {"type": "date", "value": str(value)[:10]}
  if header == "Time":
    if isinstance(value, dt.time):
      seconds = value.hour * 3600 + value.minute * 60 + value.second + value.microsecond / 1_000_000
      return seconds / 86400
    if isinstance(value, dt.datetime):
      seconds = value.hour * 3600 + value.minute * 60 + value.second + value.microsecond / 1_000_000
      return seconds / 86400
    try:
      parsed = dt.datetime.strptime(str(value), "%H:%M:%S")
      return (parsed.hour * 3600 + parsed.minute * 60 + parsed.second) / 86400
    except ValueError:
      return str(value)
  return value


all_rows = []
file_counts = []
date_counts = {}
result_counts = {}

for file_name in files:
  workbook = load_workbook(base / file_name, read_only=True, data_only=False)
  sheet = workbook["RC"]
  rows = sheet.iter_rows(values_only=True)
  source_headers = [str(value).strip() if value is not None else "" for value in next(rows)]
  header_index = {header: index for index, header in enumerate(source_headers)}
  count = 0
  for source_row in rows:
    if not any(value is not None for value in source_row):
      continue
    merged_row = []
    for header in unified_headers[:-1]:
      index = header_index.get(header)
      value = source_row[index] if index is not None and index < len(source_row) else None
      merged_row.append(normalize(value, header))
    merged_row.append(file_name)
    all_rows.append(merged_row)
    count += 1
    date_value = merged_row[0]["value"] if isinstance(merged_row[0], dict) else str(merged_row[0])
    result_value = str(merged_row[2]) if merged_row[2] is not None else "(빈값)"
    date_counts[date_value] = date_counts.get(date_value, 0) + 1
    result_counts[result_value] = result_counts.get(result_value, 0) + 1
  file_counts.append({"file": file_name, "count": count, "headers": source_headers})
  workbook.close()

payload = {
  "headers": unified_headers,
  "rows": all_rows,
  "fileCounts": file_counts,
  "dateCounts": sorted(date_counts.items()),
  "resultCounts": sorted(result_counts.items()),
}
(work / "merged_data.json").write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
print(json.dumps({
  "totalRows": len(all_rows),
  "fileCounts": file_counts,
  "dateCounts": payload["dateCounts"],
  "resultCounts": payload["resultCounts"],
}, ensure_ascii=False))
